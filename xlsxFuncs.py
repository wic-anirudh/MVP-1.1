import requests
import csv
from openpyxl import Workbook
from shariah import *
from getActiveStocks import *
from generateMasterDataArray import *
#from getBSData import *
#from getISData import *
from getData import *
#from xlsxFuncs import *
'''
def createSpreadsheet(master_data):
    print("PUTTING DATA INTO MASTER SPREADSHEET")
    # Put data in excel spreadsheet
    wb = Workbook()
    ws = wb.active

    headings = ['Ticker', 'Company Name','Industry','PE Ratio','Total Debt','Total Equity','Interest Income','Total Income']
    ws.append(headings)

    #for stock in master_data:
     #   ws.append(stock)
    completion = 0
    for row_idx, stock in enumerate(master_data, start=2):
        print(stock)
        for col_idx, value in enumerate(stock, start=1):
            print(value)
            ws.cell(row=row_idx, column=col_idx, value=value)
        print("NEXT STOCK **********************************************")
     
    wb.save("clean_stock_data.xlsx")

    print("MASTER SPREADSHEET CREATED")
'''

#from openpyxl import Workbook

def createSpreadsheet(master_data):
    print("PUTTING DATA INTO MASTER SPREADSHEET")
    wb = Workbook()
    ws = wb.active

    headings = ['Ticker', 'Company Name', 'Industry', 'PE Ratio', 'Total Debt', 'Total Equity', 'Interest Income', 'Total Income']
    ws.append(headings)
    for stock in master_data:
        ws.append(stock)

    # Iterate over rows and write data
    for row in ws.iter_rows(min_row=2, values_only=True):
        ws.append(row)

    wb.save("clean_stock_data.xlsx")
    print("MASTER SPREADSHEET CREATED")
    

def createShariahSpreadsheet():
    print("PUTTING DATA INTO SHARIAH SPREADSHEET")
    master_shariah_data = all_shariah_list()
    #print(master_shariah_data)
    # Put data in excel spreadsheet
    wb = Workbook()
    ws = wb.active

    headings = ['Ticker', 'Company Name','Industry','Prohibited Industry?','High Debt?',"High Interest Income?"]
    ws.append(headings)

    #for stock in master_shariah_data:
    #    ws.append(stock)

    for row_idx, stock in enumerate(master_shariah_data, start=2):
        for col_idx, value in enumerate(stock, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
            
    wb.save("shariah_stock_data.xlsx")

    print("SHARIAH SPREADSHEET CREATED")
