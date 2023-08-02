import pandas as pd
from typing import List
import openpyxl

def shariah_analysis(stock):
    # Get the financial data for the individual stock   -- ticker,name,industry,pe_ratio,total_debt,total_equity,interest_income,total_income
    stock_data = openpyxl.load_workbook('clean_stock_data.xlsx')
    sh = stock_data.active

    individual_stock_data = []
    for i in range(2, sh.max_row+1):
        if stock==sh.cell(row=i,column=1).value:
            for j in range(1, sh.max_column+1):
                individual_stock_data.append(sh.cell(row=i,column=j).value)
            
    #print(individual_stock_data)
   
    # Initialize a dictionary to store the results of the Shariah analysis
    results = [0,0,0]
   
    # Check if the stock is involved in any prohibited industries (e.g. gambling, alcohol, tobacco)  -- Could be more
    prohibited_industries = ['gambling', 'alcohol', 'tobacco']
    results[0] = any(industry in individual_stock_data[2] for industry in prohibited_industries)
   
    # Check if the stock has a high level of debt
    if(float(individual_stock_data[5])!=0):
        debt_to_equity_ratio = float(individual_stock_data[4]) / float(individual_stock_data[5])
        results[1] = debt_to_equity_ratio > 0.5
    else:
        results[1] = False
   
    # Check if the stock has a high level of interest-based income
    if(float(individual_stock_data[7])!=0):
        interest_income = float(individual_stock_data[6]) / float(individual_stock_data[7])
        results[2] = interest_income > 0.1
    else:
        results[2] = False
   
    # Return the results of the Shariah analysis
    return results


def all_shariah_list():
    stock_data = openpyxl.load_workbook('clean_stock_data.xlsx')
    sh = stock_data.active
    master_shariah_data = []

    for i in range(2,sh.max_row+1):
        temp=[sh.cell(row=i,column=1).value,sh.cell(row=i,column=2).value,sh.cell(row=i,column=3).value]
        temp2=shariah_analysis(sh.cell(row=i,column=1).value)
        temp +=temp2
        master_shariah_data.append(temp)
        print(temp)

    return master_shariah_data
